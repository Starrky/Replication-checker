import glob
import shutil
import openpyxl
from openpyxl import load_workbook
from os import listdir
from os.path import isfile, join
import pandas as pd
import csv
from time import time
import xlsxwriter
from xlsxwriter.utility import xl_rowcol_to_cell
import os
import subprocess

start_time = time()
cwd = os.getcwd()
results_dir = f'{cwd}\\Results'
results_dir_csv = f'{results_dir}\\csv'
results_dir_xlsx = f'{results_dir}\\xlsx'

names_csv = [f for f in listdir(results_dir_csv) if isfile(join(results_dir_csv, f))]

if "Merge.xlsx" in names_csv:
    names_csv.remove("Merge.xlsx")

for file in names_csv:
    loc = f"{results_dir_csv}\\{file}"
    wb = openpyxl.Workbook()
    ws = wb.active

    with open(loc) as f:
        reader = csv.reader(f, delimiter=';')
        for row in reader:
            ws.append(row)

    file = file.removesuffix(".csv")
    wb.save(f'{results_dir_xlsx}/{file}.xlsx')

names_xlsx = [f for f in listdir(results_dir_xlsx) if isfile(join(results_dir_xlsx, f))]
dataframes = []
for file in names_xlsx:
    xlsx = f"{results_dir_xlsx}\\{file}"
    df = pd.read_excel(xlsx)
    dataframes.append(df)

Final_excel = f'{results_dir}\\Replication_status.xlsx'
dataframes_df = pd.concat(dataframes)
dataframes_df.to_excel(Final_excel, index=False)

wb = load_workbook(Final_excel)
ws = wb.active

for row in ws.iter_rows(min_row=2, max_row=3000):
    for cell in row:
        Value = ws[f'A{cell.row}'].value
        Value = str(Value)
        Value_2 = ws[f'B{cell.row}'].value
        Value_2 = str(Value_2)
        Value_3 = ws[f'C{cell.row}'].value
        Value_3 = str(Value_3)
        Value_4 = ws[f'D{cell.row}'].value
        Value_4 = str(Value_4)

        if '-' in Value or '-' in Value_3:
            ws.delete_rows(cell.row, amount=1)

        elif "rows" in Value or '-------' in Value_3:
            ws.delete_rows(cell.row, amount=1)

        elif Value == 'None':
            ws.delete_rows(cell.row, amount=1)

        elif Value_2 == 'None':
            ws.delete_rows(cell.row, amount=1)

        elif Value_3 == 'None':
            ws.delete_rows(cell.row, amount=1)

        elif 'NULL' in Value_4:
            ws[f'D{cell.row}'] = "!No replication!"

for row in ws.iter_rows(min_row=2):
    for cell in row:
        Value_5 = ws[f'E{cell.row}'].value
        Value_5 = str(Value_5)

        if "0" in Value_5:
            ws[f'E{cell.row}'] = "!No replication!"

        elif "1" in Value_5:
            ws[f'E{cell.row}'] = "OK"

wb.save(Final_excel)

df = pd.read_excel(Final_excel)
df = df.dropna()
workbook = xlsxwriter.Workbook(Final_excel)
max_row = len(df) + 1
worksheet1 = workbook.add_worksheet('Tabela')

worksheet1.add_table(f'A1:E{max_row}', {'data': df.values.tolist(),
                                        'style': 'Table Style Light 9', 'columns':
                                            [
                                                {'header': 'Country'},
                                                {'header': 'DB'},
                                                {'header': 'Store'},
                                                {'header': 'Last replication'},
                                                {'header': 'Status'}

                                            ]})

format1 = workbook.add_format({'bold': False, 'align': 'center'})
format2 = workbook.add_format({'align': 'center', 'bg_color': '#FFC7CE'})
worksheet1.conditional_format(f'C1:C{max_row}', {'type': 'duplicate',
                                                 'format': format2})

format1.set_align('Center')
worksheet1.set_row(0, 16, format1)
worksheet1.set_column('A:A', 12, format1)
worksheet1.set_column('B:B', 15, format1)
worksheet1.set_column('C:C', 10, format1)
worksheet1.set_column('D:D', 18, format1)
worksheet1.set_column('E:E', 12, format1)

workbook.close()

dest = "C:\\Users\\jmazurek\\PEPCO\\SOC - General\\Replikacja\\"
shutil.copy(Final_excel, dest)

print("Process finished --- %s seconds ---" % (time() - start_time))
